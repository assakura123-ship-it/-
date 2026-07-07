# modules/excel_template_processor.py
"""
Процессор заполнения Excel-шаблонов карт загрузки.

Поддерживает три реальных варианта шаблона (все три листа находятся в одном
файле "ШАБЛОН КАРТЫ 2025.xlsx", который поставляется вместе с проектом):

    'oil'         -> лист "п. "        - масла цехов 2 и 3 (моторные, гидравлические и т.д.)
    'concentrate' -> лист "Концентрат" - варка растворов загущающей присадки (концентратов)
    'workshop1'   -> лист "1цех"       - масла цеха 1 (структура почти как "п. ",
                                          но добавлен столбец "по мернику", из-за чего
                                          все столбцы правее "Факт, кг" сдвинуты на 1)

Разметка (координаты ячеек, объединения, формулы) была получена построчным
разбором реального файла, загруженного пользователем, поэтому метод
fill_* для каждого типа шаблона использует ТОЧНЫЕ координаты, соответствующие
настоящему бланку, а не примерную раскладку.
"""
import os
from datetime import datetime

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side


class ExcelTemplateProcessor:
    """Процессор для заполнения шаблона Excel карты загрузки.

    Параметры
    ---------
    template_type: str
        Один из ключей TEMPLATE_TYPES: 'oil' (по умолчанию), 'concentrate',
        'workshop1'. Определяет, какой лист исходного файла используется и
        какую раскладку ячеек применять при заполнении.
    template_path: str, optional
        Путь к файлу шаблона. По умолчанию ищется файл
        "ШАБЛОН КАРТЫ 2025.xlsx" в корне проекта.
    """

    SHEET_OIL = "п. "
    SHEET_CONCENTRATE = "Концентрат"
    SHEET_WORKSHOP1 = "1цех"

    TEMPLATE_TYPES = {
        'oil': SHEET_OIL,
        'concentrate': SHEET_CONCENTRATE,
        'workshop1': SHEET_WORKSHOP1,
    }

    # Человекочитаемые названия для UI (используются, например, в выпадающем
    # списке выбора шаблона при создании карты загрузки)
    TEMPLATE_LABELS = {
        'oil': 'Масла (цех 2, цех 3) — "п."',
        'concentrate': 'Концентраты (варка загустителя)',
        'workshop1': 'Масла — Цех 1',
    }

    DEFAULT_TEMPLATE_FILENAME = "ШАБЛОН КАРТЫ 2025.xlsx"

    def __init__(self, template_type='oil', template_path=None):
        if template_type not in self.TEMPLATE_TYPES:
            template_type = 'oil'
        self.template_type = template_type
        self.sheet_name = self.TEMPLATE_TYPES[template_type]

        if template_path is None:
            # Ищем бандл-файл шаблона рядом с корнем проекта (на уровень выше modules/)
            module_dir = os.path.dirname(os.path.abspath(__file__))
            project_root = os.path.dirname(module_dir)
            candidate = os.path.join(project_root, self.DEFAULT_TEMPLATE_FILENAME)
            if os.path.exists(candidate):
                template_path = candidate
            else:
                # Резервный вариант — старое имя файла (для обратной совместимости)
                template_path = os.path.join(project_root, "ШАБЛОН КАРТЫ.xlsx")

        self.template_path = template_path
        self.wb = None
        self.ws = None

    # ===================== ЗАГРУЗКА ШАБЛОНА =====================

    def load_template(self):
        """Загрузить шаблон Excel (нужный лист по self.template_type).

        Если физический файл шаблона отсутствует (например, при переносе
        проекта на новое рабочее место без файла), генерируем эквивалентную
        раскладку программно, чтобы заполнение карты загрузки продолжало
        работать без ручной подготовки внешнего файла.
        """
        if not os.path.exists(self.template_path):
            self.wb, self.ws = self._build_default_template()
            return self.wb, self.ws

        self.wb = load_workbook(self.template_path)

        if self.sheet_name in self.wb.sheetnames:
            self.ws = self.wb[self.sheet_name]
        else:
            self.ws = self.wb.active

        return self.wb, self.ws

    # ===================== FALLBACK: ШАБЛОН "С НУЛЯ" =====================

    def _build_default_template(self):
        """Сгенерировать шаблон карты загрузки "с нуля" на случай, если
        файл "ШАБЛОН КАРТЫ 2025.xlsx" недоступен на диске.

        Разметка соответствует координатам, которые используют методы
        fill_basic_info / fill_components / fill_time_parameters /
        fill_test_results / fill_signatures для выбранного self.template_type,
        поэтому заполнение данными работает идентично оригинальному файлу.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.sheet_name

        bold = Font(bold=True)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def set_cell(coord, value=None, merge=None, is_bold=True, align=True):
            if merge:
                ws.merge_cells(merge)
            cell = ws[coord]
            if value is not None:
                cell.value = value
            if is_bold:
                cell.font = bold
            if align:
                cell.alignment = center
            cell.border = border

        if self.template_type == 'concentrate':
            self._build_default_concentrate(ws, set_cell, border)
        elif self.template_type == 'workshop1':
            self._build_default_oil_like(ws, set_cell, border, shifted=True)
        else:
            self._build_default_oil_like(ws, set_cell, border, shifted=False)

        return wb, ws

    def _build_default_oil_like(self, ws, set_cell, border, shifted=False):
        """Общая раскладка для 'oil' ("п. ") и 'workshop1' ("1цех")."""
        title_range = 'A1:M1' if shifted else 'A1:L1'
        set_cell('A1', 'Технологическая карта производства масла ', merge=title_range)

        # Верхний информационный блок
        set_cell('B2', 'время перемешивания', is_bold=False)
        set_cell('C2', '1.5ч', is_bold=False)
        set_cell('B3', 'температура', is_bold=False)
        set_cell('C3', 60, is_bold=False)
        set_cell('B4', 'циркуляция', is_bold=False)
        set_cell('C4', 'да', is_bold=False)

        if not shifted:
            set_cell('E4', 'Цех')
            set_cell('F4', 'Реактор')
            set_cell('G4', 'Замес')
            set_cell('H4', 'Размер партии, кг', merge='H4:I4')
            set_cell('J4', 'Дата выдачи ')
            set_cell('E5', None, is_bold=False)
            set_cell('F5', None, is_bold=False)
            set_cell('G5', None, is_bold=False)
            set_cell('H5', None, merge='H5:I5', is_bold=False)
            set_cell('J5', '=TODAY()', is_bold=False)
            batch_label_range = 'F7:I7'
            batch_value_cell = 'J7'
        else:
            set_cell('E4', 'Цех')
            set_cell('F4', 'Реактор')
            set_cell('G4', 'Замес')
            set_cell('I4', 'Размер партии, кг', merge='I4:J4')
            set_cell('K4', 'Дата выдачи ')
            set_cell('E5', 1, is_bold=False)
            set_cell('F5', None, is_bold=False)
            set_cell('G5', None, is_bold=False)
            set_cell('I5', None, merge='I5:J5', is_bold=False)
            set_cell('K5', '=TODAY()', is_bold=False)
            batch_label_range = 'F7:J7'
            batch_value_cell = 'K7'

        set_cell('B7', 'Дата начала приготовления', merge='B7:C7')
        set_cell('D7', None, merge='D7:E7', is_bold=False)
        set_cell('F7', 'Партия №', merge=batch_label_range)
        set_cell(batch_value_cell, None, is_bold=False)

        set_cell('B8', 'Наименование компонентов', merge='B8:C8')
        set_cell('D8', 'Рецептура', merge='D8:E8')
        set_cell('D9', 'На 100%', is_bold=False)
        set_cell('E9', 'На реактор, кг', is_bold=False)

        if not shifted:
            set_cell('F8', '№ Емкости, партии присадки', merge='F8:F9')
            set_cell('G8', 'Факт, кг', merge='G8:G9')
            set_cell('H8', 'корректировка', merge='H8:H9')
            set_cell('I8', 'Темп., °С', merge='I8:I9')
            set_cell('J8', 'Продолжительность', merge='J8:K8')
            set_cell('J9', 'время начала загрузки', is_bold=False)
            set_cell('K9', 'время окончания загрузки', is_bold=False)
            set_cell('L8', 'Подпись исполнителя')
            mass_ref = '$H$5'
        else:
            set_cell('F8', '№ Емкости, партии присадки', merge='F8:F9')
            set_cell('G8', 'Факт, кг', merge='G8:G9')
            set_cell('H8', 'по мернику', merge='H8:H9')
            set_cell('I8', 'корректировка', merge='I8:I9')
            set_cell('J8', 'Темп., °С', merge='J8:J9')
            set_cell('K8', 'Продолжительность', merge='K8:L8')
            set_cell('K9', 'время начала загрузки', is_bold=False)
            set_cell('L9', 'время окончания загрузки', is_bold=False)
            set_cell('M8', 'Подпись исполнителя')
            mass_ref = '$I$5'

        # Строки компонентов 10-19
        for row in range(10, 20):
            ws[f'A{row}'] = row - 9
            ws[f'A{row}'].border = border
            ws.merge_cells(f'B{row}:C{row}')
            cols = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
            if shifted:
                cols.append('M')
            for col in cols:
                ws[f'{col}{row}'].border = border
            if row == 10:
                ws[f'D{row}'] = f'=100-{"-".join([f"D{r}" for r in range(11, 20)])}'
            ws[f'E{row}'] = f'=(D{row}/100)*{mass_ref}'

        ws.merge_cells('A20:C20')
        ws['A20'] = 'итого'
        ws['A20'].border = border
        ws['D20'] = '=SUM(D10:D19)'
        ws['E20'] = '=SUM(E10:E19)'
        ws['F20'] = 'ИТОГО'
        for col in ['A', 'D', 'E', 'F']:
            ws[f'{col}20'].border = border
            ws[f'{col}20'].font = Font(bold=True)

        # Время (строка 22)
        set_cell('B22', 'ВРЕМЯ ВКЛЮЧЕНИЯ ЦИРКУЛЯЦИИ')
        set_cell('C22', None, is_bold=False)
        set_cell('D22', 'ВРЕМЯ ВКЛЮЧЕНИЯ НАГРЕВА', merge='D22:E22')
        set_cell('F22', None, is_bold=False)
        set_cell('G22', 'ВРЕМЯ ОТБОРА ПРОБЫ НА ВЯЗКОСТЬ БАЗОВОЙ СМЕСИ', merge='G22:H22')
        set_cell('I22', None, is_bold=False)
        if not shifted:
            set_cell('J22', 'ВРЕМЯ ОТБОРА ПРОБЫ НА ГОТОВНОСТЬ, ОТКЛЮЧЕНИЕ ПЕРЕМЕШИВАНИЯ И НАГРЕВА', merge='J22:K22')
            set_cell('L22', None, is_bold=False)
        else:
            set_cell('J22', 'ВРЕМЯ ОТБОРА ПРОБЫ НА ГОТОВНОСТЬ, ОТКЛЮЧЕНИЕ ПЕРЕМЕШИВАНИЯ И НАГРЕВА', merge='J22:L22')
            set_cell('M22', None, is_bold=False)

        set_cell('B24', 'Вязкость базовой смеси расчетная  =', merge='B24:E24')
        set_cell('B25', 'Вязкость базовой смеси фактическая  =', merge='B25:E25')

        set_cell('B27', 'Результаты испытаний', merge='B27:L27')
        set_cell('B28', 'Стадия', merge='B28:B29')
        set_cell('C28', 'КВ при 100°С', merge='C28:D28')
        set_cell('E28', 'КВ при 40°С', merge='E28:F28')
        set_cell('G28', 'ИВ', merge='G28:H28')
        set_cell('I28', 'CCS ', merge='I28:J28')
        set_cell('K28', 'Тзаст', merge='K28:L28')
        for col, text in zip('CDEFGHIJKL', ['норма', 'факт'] * 5):
            set_cell(f'{col}29', text, is_bold=False)

        set_cell('B30', 'готовое масло', is_bold=True, align=False)
        set_cell('C30', None, merge='C30:C32', is_bold=False)
        set_cell('D30', None, is_bold=False)
        set_cell('E30', 'не норм, определение обязательно', merge='E30:E32', is_bold=False)
        set_cell('F30', None, is_bold=False)
        set_cell('G30', None, merge='G30:G32', is_bold=False)
        set_cell('H30', None, is_bold=False)
        set_cell('I30', None, merge='I30:I32', is_bold=False)
        set_cell('J30', None, is_bold=False)
        set_cell('K30', None, merge='K30:K32', is_bold=False)
        set_cell('L30', None, is_bold=False)
        set_cell('B31', 'корректировка №1', is_bold=True, align=False)
        set_cell('B32', 'корректировка №2', is_bold=True, align=False)
        for row in (31, 32):
            for col in ('D', 'F', 'H', 'J', 'L'):
                set_cell(f'{col}{row}', None, is_bold=False)

        if not shifted:
            set_cell('B33', 'Соответствие продукции требуемым нормам ТУ, ГОСТ (да/нет)', merge='B33:H33', align=False)
            set_cell('I33', None, is_bold=False)
            set_cell('J33', 'Подпись лаборанта', merge='J33:K33')
            set_cell('L33', None, is_bold=False)
            set_cell('B35', 'СКАЧАЛИ В Е№')
            set_cell('C35', None, is_bold=False)
            set_cell('F35', 'Аппаратчик')
            set_cell('G35', None, merge='G35:H35', is_bold=False)
            set_cell('I35', 'дата')
            set_cell('J35', None, is_bold=False)
        else:
            set_cell('B33', 'Соответствие продукции требуемым нормам ТУ, ГОСТ (да/нет)', merge='B33:I33', align=False)
            set_cell('J33', 'Подпись лаборанта', merge='J33:K33')
            set_cell('L33', None, is_bold=False)
            set_cell('B35', 'СКАЧАЛИ В Е№')
            set_cell('C35', None, is_bold=False)
            set_cell('F35', 'Аппаратчик')
            set_cell('G35', None, merge='G35:I35', is_bold=False)
            set_cell('J35', 'дата')
            set_cell('K35', None, is_bold=False)

    def _build_default_concentrate(self, ws, set_cell, border):
        """Раскладка для листа 'Концентрат'."""
        set_cell('A1', 'Технологическая карта производства раствора загущающей присадки', merge='A1:L1')

        set_cell('B2', 'время перемешивания', is_bold=False)
        set_cell('C2', '6-10ч до растворения', is_bold=False)
        set_cell('E2', 'Загуститель ', merge='E2:F2')
        set_cell('G2', '% ввода', merge='G2:H2')
        set_cell('I2', 'Растворитель', merge='I2:J2')

        set_cell('B3', 'температура', is_bold=False)
        set_cell('C3', '120-130', is_bold=False)
        set_cell('E3', None, merge='E3:F3', is_bold=False)
        set_cell('G3', None, merge='G3:H3', is_bold=False)
        set_cell('I3', None, merge='I3:J3', is_bold=False)

        set_cell('B4', 'циркуляция', is_bold=False)
        set_cell('C4', 'да', is_bold=False)
        set_cell('E4', 'Цех')
        set_cell('F4', 'Реактор')
        set_cell('G4', 'Замес')
        set_cell('H4', 'Размер партии, кг', merge='H4:I4')
        set_cell('J4', 'Дата выдачи ')
        set_cell('E5', None, is_bold=False)
        set_cell('F5', None, is_bold=False)
        set_cell('G5', None, is_bold=False)
        set_cell('H5', None, merge='H5:I5', is_bold=False)
        set_cell('J5', '=TODAY()', is_bold=False)

        set_cell('B7', 'Дата начала приготовления', merge='B7:C7')
        set_cell('D7', None, merge='D7:E7', is_bold=False)
        set_cell('F7', 'Партия №', merge='F7:I7')
        set_cell('J7', None, is_bold=False)

        set_cell('B8', 'Наименование компонентов', merge='B8:C8')
        set_cell('D8', 'Рецептура', merge='D8:E8')
        set_cell('D9', 'На 100%', is_bold=False)
        set_cell('E9', 'На реактор, кг', is_bold=False)
        set_cell('F8', '№ Емкости, партии присадки', merge='F8:F9')
        set_cell('G8', 'Факт, кг', merge='G8:G9')
        set_cell('H8', 'корректировка', merge='H8:H9')
        set_cell('I8', 'Темп., °С', merge='I8:I9')
        set_cell('J8', 'Продолжительность', merge='J8:K8')
        set_cell('J9', 'время начала загрузки', is_bold=False)
        set_cell('K9', 'время окончания загрузки', is_bold=False)
        set_cell('L8', 'Подпись исполнителя')

        for row in range(10, 13):
            ws[f'A{row}'] = row - 9
            ws[f'A{row}'].border = border
            ws.merge_cells(f'B{row}:C{row}')
            for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                ws[f'{col}{row}'].border = border
            ws[f'E{row}'] = f'=(D{row}/100)*$H$5'

        ws['B10'] = '=I3'
        ws['D10'] = '=100-D11-D12'
        ws['B11'] = '=E3'
        ws['D11'] = '=G3'

        ws.merge_cells('A13:C13')
        ws['A13'] = 'итого'
        ws['A13'].border = border
        ws['D13'] = '=SUM(D10:D12)'
        ws['E13'] = '=SUM(E10:E12)'
        ws['F13'] = 'ИТОГО'
        for col in ['A', 'D', 'E', 'F']:
            ws[f'{col}13'].border = border
            ws[f'{col}13'].font = Font(bold=True)

        set_cell('B15', 'ВРЕМЯ ВКЛЮЧЕНИЯ ЦИРКУЛЯЦИИ')
        set_cell('C15', None, is_bold=False)
        set_cell('D15', 'ВРЕМЯ ВКЛЮЧЕНИЯ НАГРЕВА', merge='D15:E15')
        set_cell('F15', None, is_bold=False)
        set_cell('G15', 'ВРЕМЯ ОТБОРА ПРОМЕЖУТОЧНОЙ ПРОБЫ', merge='G15:H15')
        set_cell('I15', None, is_bold=False)
        set_cell('J15', 'ВРЕМЯ ОТБОРА ПРОБЫ НА ГОТОВНОСТЬ, ОТКЛЮЧЕНИЕ ПЕРЕМЕШИВАНИЯ И НАГРЕВА', merge='J15:K15')
        set_cell('L15', None, is_bold=False)

        set_cell('B17', 'Результаты испытаний')
        set_cell('C17', 'КВ при 100°С', merge='C17:D18')
        set_cell('E17', 'Для использования в маслах:', merge='E17:L18')
        set_cell('B18', 'Стадия', merge='B18:B19')
        set_cell('C19', 'норма', is_bold=False)
        set_cell('D19', 'факт', is_bold=False)
        set_cell('E19', None, merge='E19:L19', is_bold=False)

        set_cell('B20', 'готовое масло', is_bold=True, align=False)
        set_cell('C20', None, merge='C20:C22', is_bold=False)
        set_cell('D20', None, is_bold=False)
        set_cell('E20', None, merge='E20:L20', is_bold=False)
        set_cell('B21', 'корректировка №1', is_bold=True, align=False)
        set_cell('D21', None, is_bold=False)
        set_cell('E21', None, merge='E21:L21', is_bold=False)
        set_cell('B22', 'корректировка №2', is_bold=True, align=False)
        set_cell('D22', None, is_bold=False)
        set_cell('E22', None, merge='E22:L22', is_bold=False)

        set_cell('B24', 'Остаток слит в куб №')
        set_cell('C24', None, is_bold=False)
        set_cell('F24', 'Аппаратчик')
        set_cell('G24', None, merge='G24:H24', is_bold=False)
        set_cell('I24', 'дата')
        set_cell('J24', None, is_bold=False)

    # ===================== ЗАПОЛНЕНИЕ ДАННЫМИ =====================

    def fill_basic_info(self, data):
        """Заполнить базовую информацию (шапка карты)."""
        if self.template_type == 'concentrate':
            self.ws['A1'] = (
                "Технологическая карта производства раствора загущающей присадки"
                + (f" {data['product_name']}" if data.get('product_name') else "")
            )
        else:
            self.ws['A1'] = f"Технологическая карта производства {data.get('product_name', 'масла')}"

        # Цех / Реактор / Замес / Размер партии / Дата выдачи
        self.ws['E5'] = data.get('workshop', '')
        self.ws['F5'] = data.get('reactor', '')
        self.ws['G5'] = data.get('batch_type', 'Замес')

        if self.template_type == 'workshop1':
            self.ws['I5'] = data.get('batch_quantity', '')
        else:
            self.ws['H5'] = data.get('batch_quantity', '')

        if data.get('issue_date'):
            if self.template_type == 'workshop1':
                self.ws['K5'] = data['issue_date']
            else:
                self.ws['J5'] = data['issue_date']

        # Дата начала приготовления и номер партии
        self.ws['D7'] = data.get('start_date', '')
        if self.template_type == 'workshop1':
            self.ws['K7'] = data.get('batch_number', '')
        else:
            self.ws['J7'] = data.get('batch_number', '')

        # Для листа "Концентрат" дополнительно заполняем блок
        # Загуститель / % ввода / Растворитель (используется формулами
        # в таблице компонентов B10/D10/B11/D11)
        if self.template_type == 'concentrate':
            if data.get('solvent_name') is not None:
                self.ws['I3'] = data.get('solvent_name', '')
            if data.get('thickener_name') is not None:
                self.ws['E3'] = data.get('thickener_name', '')
            if data.get('concentrate_percentage') is not None:
                self.ws['G3'] = data.get('concentrate_percentage', '')

    def fill_components(self, components):
        """Заполнить таблицу компонентов.

        Для 'oil' и 'workshop1' — до 10 строк (10-19), проценты второй и
        последующих строк вписываются напрямую, первая строка (10) остаётся
        формулой "остаток до 100%" — как в оригинальном шаблоне.

        Для 'concentrate' ожидается порядок компонентов:
            components[0] -> Растворитель (имя автоматически берётся из I3,
                              процент — формула остатка, напрямую не пишем)
            components[1] -> Загуститель (его имя/процент пишутся в поля
                              E3/G3 в fill_basic_info; сюда не дублируем)
            components[2] -> опциональный третий компонент (пишется напрямую
                              в строку 12)
        Во всех случаях операционные поля (номер ёмкости, факт, коррекция,
        температура, время, подпись) заполняются для каждой присутствующей
        строки.
        """
        if self.template_type == 'concentrate':
            self._fill_components_concentrate(components)
        else:
            self._fill_components_oil_like(components, shifted=(self.template_type == 'workshop1'))

    def _fill_components_oil_like(self, components, shifted=False):
        start_row = 10
        max_rows = 10

        if not shifted:
            col_map = {
                'container': 'F', 'fact': 'G', 'correction': 'H',
                'temperature': 'I', 'start_time': 'J', 'end_time': 'K', 'signature': 'L',
            }
        else:
            col_map = {
                'container': 'F', 'fact': 'G', 'gauge': 'H', 'correction': 'I',
                'temperature': 'J', 'start_time': 'K', 'end_time': 'L', 'signature': 'M',
            }

        for i, comp in enumerate(components):
            row = start_row + i
            if row >= start_row + max_rows:
                break

            self.ws[f'B{row}'] = comp.get('name', '')

            # Первая строка (10) в реальном шаблоне — формула остатка до 100%.
            # Не перезаписываем её, чтобы сохранить встроенную проверку суммы.
            if row != start_row:
                self.ws[f'D{row}'] = comp.get('percentage', 0)

            self.ws[f'{col_map["container"]}{row}'] = comp.get('container_number', '')
            self.ws[f'{col_map["fact"]}{row}'] = comp.get('actual_mass', '')
            if shifted:
                self.ws[f'{col_map["gauge"]}{row}'] = comp.get('gauge_reading', '')
            self.ws[f'{col_map["correction"]}{row}'] = comp.get('correction', '')
            self.ws[f'{col_map["temperature"]}{row}'] = comp.get('temperature', '')
            self.ws[f'{col_map["start_time"]}{row}'] = comp.get('start_time', '')
            self.ws[f'{col_map["end_time"]}{row}'] = comp.get('end_time', '')
            self.ws[f'{col_map["signature"]}{row}'] = comp.get('signature', '')

    def _fill_components_concentrate(self, components):
        col_map = {
            'container': 'F', 'fact': 'G', 'correction': 'H',
            'temperature': 'I', 'start_time': 'J', 'end_time': 'K', 'signature': 'L',
        }

        for i, comp in enumerate(components[:3]):
            row = 10 + i
            # Имена/проценты строк 10 и 11 приходят автоматически по формулам
            # (из E3/G3/I3, заполненных в fill_basic_info). Прямо пишем
            # только для третьего (опционального) компонента — строка 12.
            if i == 2:
                self.ws[f'B{row}'] = comp.get('name', '')
                self.ws[f'D{row}'] = comp.get('percentage', 0)

            self.ws[f'{col_map["container"]}{row}'] = comp.get('container_number', '')
            self.ws[f'{col_map["fact"]}{row}'] = comp.get('actual_mass', '')
            self.ws[f'{col_map["correction"]}{row}'] = comp.get('correction', '')
            self.ws[f'{col_map["temperature"]}{row}'] = comp.get('temperature', '')
            self.ws[f'{col_map["start_time"]}{row}'] = comp.get('start_time', '')
            self.ws[f'{col_map["end_time"]}{row}'] = comp.get('end_time', '')
            self.ws[f'{col_map["signature"]}{row}'] = comp.get('signature', '')

    def fill_time_parameters(self, time_data):
        """Заполнить временные параметры."""
        if self.template_type == 'concentrate':
            self.ws['C15'] = time_data.get('circulation_start', '')
            self.ws['F15'] = time_data.get('heating_start', '')
            self.ws['I15'] = time_data.get('intermediate_sample_time',
                                            time_data.get('viscosity_test_time', ''))
            self.ws['L15'] = time_data.get('readiness_test_time', '')
            return

        self.ws['C22'] = time_data.get('circulation_start', '')
        self.ws['F22'] = time_data.get('heating_start', '')
        self.ws['I22'] = time_data.get('viscosity_test_time', '')

        if self.template_type == 'workshop1':
            self.ws['M22'] = time_data.get('readiness_test_time', '')
        else:
            self.ws['L22'] = time_data.get('readiness_test_time', '')

        calc_v = time_data.get('calculated_viscosity', '')
        actual_v = time_data.get('actual_viscosity', '')
        self.ws['B24'] = f"Вязкость базовой смеси расчетная  = {calc_v}"
        self.ws['B25'] = f"Вязкость базовой смеси фактическая  = {actual_v}"

    def fill_test_results(self, test_results):
        """Заполнить результаты испытаний."""
        if self.template_type == 'concentrate':
            self.ws['C20'] = test_results.get('kv_100_norm', '')
            self.ws['D20'] = test_results.get('kv_100_actual', '')
            self.ws['D21'] = test_results.get('kv_100_actual_c1', '')
            self.ws['D22'] = test_results.get('kv_100_actual_c2', '')
            if test_results.get('usage_note') is not None:
                self.ws['E20'] = test_results.get('usage_note', '')
            return

        # 'oil' / 'workshop1' — общая структура таблицы испытаний (строки 30-32)
        self.ws['C30'] = test_results.get('kv_100_norm', '')
        self.ws['D30'] = test_results.get('kv_100_actual', '')
        self.ws['F30'] = test_results.get('kv_40_actual', '')
        self.ws['G30'] = test_results.get('iv_norm', '')
        self.ws['H30'] = test_results.get('iv_actual', '')
        self.ws['I30'] = test_results.get('ccs_norm', '')
        self.ws['J30'] = test_results.get('ccs_actual', '')
        self.ws['K30'] = test_results.get('tzast_norm', '')
        self.ws['L30'] = test_results.get('tzast_actual', '')

        self.ws['D31'] = test_results.get('kv_100_actual_c1', '')
        self.ws['F31'] = test_results.get('kv_40_actual_c1', '')
        self.ws['H31'] = test_results.get('iv_actual_c1', '')
        self.ws['J31'] = test_results.get('ccs_actual_c1', '')
        self.ws['L31'] = test_results.get('tzast_actual_c1', '')

        self.ws['D32'] = test_results.get('kv_100_actual_c2', '')
        self.ws['F32'] = test_results.get('kv_40_actual_c2', '')
        self.ws['H32'] = test_results.get('iv_actual_c2', '')
        self.ws['J32'] = test_results.get('ccs_actual_c2', '')
        self.ws['L32'] = test_results.get('tzast_actual_c2', '')

    def fill_signatures(self, signatures):
        """Заполнить подписи."""
        if self.template_type == 'concentrate':
            self.ws['C24'] = signatures.get('residue_container_number', '')
            self.ws['G24'] = signatures.get('operator', '')
            self.ws['J24'] = signatures.get('completion_date', '')
            return

        if self.template_type == 'oil':
            self.ws['I33'] = signatures.get('compliance', '')
        # Для 'workshop1' в реальном шаблоне нет отдельной ячейки для
        # "да/нет" (объединена вместе с текстом заголовка) — пропускаем.

        self.ws['L33'] = signatures.get('lab_technician', '')
        self.ws['G35'] = signatures.get('operator', '')

        if self.template_type == 'workshop1':
            self.ws['K35'] = signatures.get('completion_date', '')
        else:
            self.ws['J35'] = signatures.get('completion_date', '')

    def save(self, output_path):
        """Сохранить заполненный шаблон"""
        if self.wb:
            self.wb.save(output_path)

    def create_from_card_data(self, card_data, output_path):
        """Создать файл из данных карты.

        card_data может содержать ключ 'template_type' ('oil' | 'concentrate'
        | 'workshop1'), который переопределит тип, заданный в конструкторе
        (удобно, когда процессор создаётся один раз, а тип определяется по
        выбранной номенклатурной группе продукта).
        """
        try:
            requested_type = card_data.get('template_type')
            if requested_type and requested_type in self.TEMPLATE_TYPES:
                self.template_type = requested_type
                self.sheet_name = self.TEMPLATE_TYPES[requested_type]

            # Загружаем шаблон
            self.load_template()

            # Заполняем базовую информацию
            self.fill_basic_info({
                'product_name': card_data.get('product_name', ''),
                'workshop': card_data.get('workshop', '1 цех'),
                'reactor': card_data.get('reactor', 'Р-1'),
                'batch_type': card_data.get('batch_type', 'Замес'),
                'batch_quantity': card_data.get('batch_quantity', 1000),
                'issue_date': card_data.get('issue_date', datetime.now().strftime('%d.%m.%Y')),
                'batch_number': card_data.get('batch_number', f"П-{datetime.now().strftime('%Y%m%d')}"),
                'start_date': card_data.get('start_date', datetime.now().strftime('%d.%m.%Y')),
                'solvent_name': card_data.get('solvent_name'),
                'thickener_name': card_data.get('thickener_name'),
                'concentrate_percentage': card_data.get('concentrate_percentage'),
            })

            # Заполняем компоненты
            self.fill_components(card_data.get('components', []))

            # Заполняем временные параметры
            self.fill_time_parameters(card_data.get('time_data', {}))

            # Заполняем результаты испытаний
            self.fill_test_results(card_data.get('test_results', {}))

            # Заполняем подписи
            self.fill_signatures(card_data.get('signatures', {}))

            # Сохраняем
            self.save(output_path)
            return True

        except Exception as e:
            print(f"Ошибка создания Excel: {e}")
            return False
